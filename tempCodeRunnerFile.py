# app_spread.py · v8 (Estável - v2 + Funcionalidades Corrigidas)
# --------------------------------------------------------------------
# • Baseada na v2 para garantir estabilidade e funcionalidade completa.
# • Lógica de atualização de fórmulas (incluindo complexas) restaurada.
# • Funções de DRE e DFC restauradas e funcionando.
# • Destaque com 2 cores (Verde/Amarelo) implementado de forma segura.
# • Otimização de performance mantida.
# • Relatório detalhado mantido.
# • Interface em Português e simplificada conforme solicitado.
# --------------------------------------------------------------------
import logging
import re
from pathlib import Path
from tkinter import filedialog
from typing import Callable, Dict, List, Tuple, Set

import customtkinter as ctk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import (
    column_index_from_string as col2idx,
    get_column_letter as idx2col,
)

try:
    import xlwings as xw
    XLWINGS = True
except ImportError:
    XLWINGS = False


def normaliza_num(v) -> int | None:
    if v is None or (isinstance(v, float) and pd.isna(v)): return None
    if isinstance(v, (int, float)): return int(v)
    if isinstance(v, str):
        s = v.strip().replace(".", "").replace(",", "")
        if re.fullmatch(r"-?\d+", s): return int(s)
    return None


def periodos(p: str) -> Tuple[str, str, str, bool]:
    p = p.upper().strip()
    if re.fullmatch(r"\d{4}", p):
        a = int(p)
        return str(a), str(a - 1), str(a - 2), False
    m = re.fullmatch(r"([1-4])T(\d{2})", p)
    if not m: raise ValueError("Período deve ser AAAA ou nTAA (ex.: 2024 ou 1T25).")
    tri, aa = int(m.group(1)), int(m.group(2))
    f = lambda y: f"{tri}T{y:02d}"
    return f(aa), f(aa - 1), f(aa - 2), True


def col_txt_to_idx(txt: str) -> int:
    txt = txt.strip().upper()
    return int(txt) if txt.isdigit() else col2idx(txt) - 1


def prepara_origem(
    path: Path, tipo: str, atual: str, ant: str, ant2: str, is_trim: bool,
    out_dir: Path | None
) -> Path:
    dst_dir = out_dir or path.parent
    dst_dir.mkdir(parents=True, exist_ok=True)
    dst_path = dst_dir / f"{path.stem}_tratado{path.suffix}"
    sheet_map = {
        "consolidado": {
            "DF Cons Ativo": "cons ativos", "DF Cons Passivo": "cons passivos",
            "DF Cons Resultado Periodo": "cons DRE", "DF Cons Fluxo de Caixa": "cons DFC",
            "DF Cons DMPL Ultimo": "cons DMPL",
        },
        "individual": {
            "DF Ind Ativo": "ind ativos", "DF Ind Passivo": "ind passivos",
            "DF Ind Resultado Periodo": "ind DRE", "DF Ind Fluxo de Caixa": "ind DFC",
            "DF Ind DMPL Ultimo": "ind DMPL",
        },
    }[tipo]
    H_ANO = ("valor ultimo exercicio", "valor penultimo exercicio", "valor antepenultimo exercicio")
    H_TRI_AP = ("valor trimestre atual", "valor exercicio anterior")
    H_TRI_RES = ("valor acumulado atual exercicio", "valor acumulado exercicio anterior")

    def ren_factory(sheet_orig: str) -> Callable[[str], str]:
        low = sheet_orig.lower()
        is_ap = any(k in low for k in ("ativo", "passivo"))
        is_res = "resultado" in low
        def ren(c: str) -> str:
            c_low = c.lower().strip()
            if is_trim and is_ap:
                if c_low.startswith(H_TRI_AP[0]): return atual
                if c_low.startswith(H_TRI_AP[1]): return ant
            elif is_trim and is_res:
                if c_low.startswith(H_TRI_RES[0]): return atual
                if c_low.startswith(H_TRI_RES[1]): return ant
            if c_low.startswith(H_ANO[0]): return atual
            if c_low.startswith(H_ANO[1]): return ant
            if c_low.startswith(H_ANO[2]): return ant2
            return c
        return ren

    with pd.ExcelWriter(dst_path, engine="openpyxl") as wr:
        with pd.ExcelFile(path, engine="openpyxl") as xls:
            for orig, novo in sheet_map.items():
                if orig in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=orig)
                    df = df.rename(columns=ren_factory(orig))
                    df.to_excel(wr, sheet_name=novo, index=False)
    
    wb = load_workbook(dst_path)
    if wb.sheetnames: wb.active.sheet_state = "visible"
    wb.save(dst_path)
    return dst_path


def shift_formula(f: str, delta: int) -> str:
    pat = re.compile(r"(?<![A-Za-z0-9_])(\$?)([A-Za-z]{1,3})(?=\$?\d|:)", flags=re.I)
    def repl(m: re.Match) -> str:
        if m.group(1) is None: return m.group(0)
        abs_, col = m.group(1), m.group(2)
        try:
            new_col = idx2col(col2idx(col.upper()) + delta)
            return f"{abs_}{new_col}"
        except ValueError:
            return m.group(0)
    return pat.sub(repl, f)


def valor_corresp(abas: Dict[str, pd.DataFrame], n: int, prev: str, curr: str) -> int | None:
    for df in abas.values():
        if prev in df.columns and curr in df.columns:
            hit = df.loc[df[prev].apply(normaliza_num) == n, curr]
            if not hit.empty: return normaliza_num(hit.iloc[0])
    return None

def adjust_complex_formula(
    formula: str, delta: int, map_number: Callable, report: Dict, r: int,
    used_vals_changed: Set[int], used_vals_new: Set[int]
) -> str:
    num_pat = re.compile(r"(?<![A-Za-z])[-+]?\d[\d\.,]*")
    formula_shifted = shift_formula(formula, delta)

    def repl(m: re.Match) -> str:
        n_ant = normaliza_num(m.group(0))
        if n_ant is None: return m.group(0)
        
        n_atual = map_number(n_ant)
        if n_atual is not None:
            if n_ant != 0:
                used_vals_changed.add(n_atual)
                report["alterados"].append({"ant": n_ant, "atual": n_atual, "linha": r})
            else:
                used_vals_new.add(n_atual)
                report["novos"].append({"atual": n_atual, "linha": r})
            return str(n_atual)
        
        if n_ant != 0: report["ignorados"].append({"valor": n_ant, "linha": r})
        return m.group(0)

    return num_pat.sub(repl, formula_shifted)


def atualizar_ws(
    get_val, set_val, abas: Dict[str, pd.DataFrame], src_idx: int, dst_idx: int,
    atual: str, ant: str, start_row: int
) -> Tuple[Dict, Set[int], Set[int]]:
    report = {"alterados": [], "novos": [], "ignorados": []}
    used_vals_changed, used_vals_new = set(), set()
    c_src, c_dst, delta = src_idx + 1, dst_idx + 1, dst_idx - src_idx
    
    empty_streak, r = 0, start_row
    while empty_streak < 30 and r <= 1_048_576:
        v = get_val(r, c_src)
        if v in (None, ""):
            empty_streak += 1; r += 1; continue
        empty_streak = 0
        
        destino, wrote = v, False
        if isinstance(v, str) and v.startswith("="):
            map_func = lambda n: valor_corresp(abas, n, ant, atual)
            formula_final = adjust_complex_formula(
                v, delta, map_func, report, r, used_vals_changed, used_vals_new
            )
            if formula_final != shift_formula(v, delta):
                destino, wrote = formula_final, True
        
        elif (n_ant := normaliza_num(v)) is not None:
            n_atual = valor_corresp(abas, n_ant, ant, atual)
            if n_atual is not None:
                destino, wrote = n_atual, True
                if n_ant != 0:
                    used_vals_changed.add(n_atual)
                    report["alterados"].append({"ant": n_ant, "atual": n_atual, "linha": r})
                else:
                    used_vals_new.add(n_atual)
                    report["novos"].append({"atual": n_atual, "linha": r})
            elif n_ant != 0:
                report["ignorados"].append({"valor": n_ant, "linha": r})
        else:
            destino, wrote = v, True

        if wrote:
            try: set_val(r, c_dst, destino)
            except Exception: pass
        r += 1

    return report, used_vals_changed, used_vals_new


DRE_MAP = {0: "Receita de Venda de Bens e/ou Serviços", 13: "Custo dos Bens e/ou Serviços Vendidos", 23: "Despesas Gerais e Administrativas", 25: "Despesas com Vendas", 26: "Outras Receitas Operacionais", 27: "Outras Despesas Operacionais", 29: "Despesas Financeiras", 30: "Receitas Financeiras", 41: "Resultado de Equivalência Patrimonial", 43: "Imposto de Renda e Contribuição Social sobre o Lucro"}

def aplicar_dre_manual(df_dre, sheet, col_dst_1based, dre_start, col_valor, is_xlwings):
    if df_dre is None: return
    for offset, desc in DRE_MAP.items():
        linha = dre_start + offset
        try:
            raw_val = df_dre.loc[df_dre["Descricao Conta"].str.strip() == desc, col_valor].iloc[0]
            valor = normaliza_num(raw_val) or raw_val
            if is_xlwings: sheet.cells(linha, col_dst_1based).value = valor
            else: sheet.cell(linha, col_dst_1based, value=valor)
        except (IndexError, KeyError): continue

def inserir_depreciacao_dfc(df_dfc, sheet, col_dst_1based, linha, col_valor, is_xlwings) -> int | None:
    if df_dfc is None or col_valor not in df_dfc.columns: return None
    desc = df_dfc["Descricao Conta"].astype(str)
    mask = desc.str.contains("deprecia|amortiza", case=False, na=False)
    try:
        raw_val = df_dfc.loc[mask, col_valor].iloc[0]
        num_val = normaliza_num(raw_val)
        valor = -abs(num_val) if num_val is not None else f"-{str(raw_val).lstrip('+-')}"
        if is_xlwings: sheet.cells(linha, col_dst_1based).value = valor
        else: sheet.cell(linha, col_dst_1based, value=valor)
        return normaliza_num(valor)
    except (IndexError, KeyError): return None

def destacar_inseridos(
    orig_tratada: Path, used_vals_changed: Set[int], used_vals_new: Set[int],
    atual: str, prefer_xlwings: bool = True
):
    if not used_vals_changed and not used_vals_new: return
    if prefer_xlwings and XLWINGS:
        try:
            with xw.App(visible=False) as app:
                with app.books.open(str(orig_tratada)) as wb:
                    for sht in wb.sheets:
                        if not hasattr(sht.api, 'Visible') or sht.api.Visible != -1: continue # xlSheetVisible
                        headers = sht.range("A1").expand("right").value
                        if not headers: continue
                        headers = [headers] if not isinstance(headers, list) else headers
                        atual_cols = [i + 1 for i, h in enumerate(headers) if str(h).strip() == atual]
                        for c_idx in atual_cols:
                            last_row = sht.cells.last_cell.row
                            if last_row <= 1: continue
                            rng = sht.range((2, c_idx), (last_row, c_idx))
                            vals = rng.value
                            to_green, to_yellow = [], []
                            vals = [vals] if not isinstance(vals, list) else vals
                            for i, val in enumerate(vals or []):
                                norm_val = normaliza_num(val)
                                if norm_val in used_vals_changed: to_green.append(rng[i])
                                elif norm_val in used_vals_new: to_yellow.append(rng[i])
                            for cell in to_green: cell.color, cell.font.bold = (204, 255, 204), True
                            for cell in to_yellow: cell.color, cell.font.bold = (255, 255, 153), True
                    wb.save()
            return
        except Exception as e:
            logging.warning(f"Falha ao destacar com xlwings: {e}. Usando openpyxl.")

    fill_changed, fill_new, bold = PatternFill("solid", fgColor="CCFFCC"), PatternFill("solid", fgColor="FFFF99"), Font(bold=True)
    wb = load_workbook(orig_tratada, keep_vba=orig_tratada.suffix.lower() == ".xlsm")
    for ws in wb.worksheets:
        atual_cols = [c.column for c in ws[1] if str(c.value).strip() == atual]
        for row in ws.iter_rows(min_row=2):
            for c_idx in atual_cols:
                cell = row[c_idx - 1]
                val = normaliza_num(cell.value)
                if val in used_vals_changed: cell.fill, cell.font = fill_changed, bold
                elif val in used_vals_new: cell.fill, cell.font = fill_new, bold
    wb.save(orig_tratada)

def processar(
    ori: Path, spr: Path, tipo: str, periodo: str, src_txt: str, dst_txt: str,
    start_row: int, dre_start: int, out_dir: Path | None = None,
    log: Callable[[str], None] = print
) -> Dict:
    src_idx, dst_idx = col_txt_to_idx(src_txt), col_txt_to_idx(dst_txt)
    atual, ant, ant2, is_trim = periodos(periodo)
    orig_path = prepara_origem(ori, tipo, atual, ant, ant2, is_trim, out_dir)
    abas = pd.read_excel(orig_path, sheet_name=None, engine="openpyxl")
    dre_df = abas.get(f"{'cons' if tipo == 'consolidado' else 'ind'} DRE")
    dfc_df = abas.get(f"{'cons' if tipo == 'consolidado' else 'ind'} DFC")

    report, used_changed, used_new, processed_ok = {}, set(), set(), False
    
    if XLWINGS and spr.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            with xw.App(visible=False) as app:
                with app.books.open(str(spr)) as wb:
                    sht = wb.sheets[0]
                    get_val = lambda r, c: sht.cells(r, c).formula
                    set_val = lambda r, c, v: setattr(sht.cells(r, c), 'value', v)
                    report, used_changed, used_new = atualizar_ws(get_val, set_val, abas, src_idx, dst_idx, atual, ant, start_row)
                    if is_trim: aplicar_dre_manual(dre_df, sht, dst_idx + 1, dre_start, atual, True)
                    if (v_dfc := inserir_depreciacao_dfc(dfc_df, sht, dst_idx + 1, 199, atual, True)):
                        used_changed.add(v_dfc)
                    wb.app.calculate(); wb.save()
                processed_ok = True
        except Exception as e: log(f"xlwings falhou: {e}. Usando fallback.")

    if not processed_ok:
        is_xlsm = spr.suffix.lower() == ".xlsm"
        wb = load_workbook(spr, keep_vba=is_xlsm)
        ws = wb.active
        get_val = lambda r, c: ws.cell(r, c).value
        set_val = lambda r, c, v: setattr(ws.cell(r, c), 'value', v)
        report, used_changed, used_new = atualizar_ws(get_val, set_val, abas, src_idx, dst_idx, atual, ant, start_row)
        if is_trim: aplicar_dre_manual(dre_df, ws, dst_idx + 1, dre_start, atual, False)
        if (v_dfc := inserir_depreciacao_dfc(dfc_df, ws, dst_idx + 1, 199, atual, False)):
            used_changed.add(v_dfc)
        out_name = f"{spr.stem} {atual}{'.xlsm' if is_xlsm else '.xlsx'}"
        spr_out = spr.with_name(out_name)
        wb.save(spr_out)
        log(f"Spread atualizado salvo em: {spr_out}")

    destacar_inseridos(orig_path, used_changed, used_new, atual, XLWINGS)
    log(f"Origem tratada salva e destacada em: {orig_path}")
    return report

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Atualizador de Spread")
        self.grid_columnconfigure((0, 1), weight=1)
        self.var_ori = ctk.StringVar()
        self._campo_arquivo("Arquivo Origem", 0, self.var_ori)
        self.var_spr = ctk.StringVar()
        self._campo_arquivo("Arquivo Spread", 1, self.var_spr)
        self.var_tipo = ctk.StringVar(value="consolidado")
        ctk.CTkLabel(self, text="Tipo").grid(row=2, column=0, sticky="w", padx=4, pady=(5,0))
        ctk.CTkOptionMenu(self, variable=self.var_tipo, values=["consolidado", "individual"]).grid(row=2, column=1, sticky="ew", padx=4, pady=(5,0))
        self.var_per = ctk.StringVar()
        self._campo_txt("Período (Ex: 2024 ou 1T25)", 3, self.var_per)
        self.var_src = ctk.StringVar(value="A")
        self._campo_txt("Coluna Origem (A ou 0…)", 4, self.var_src, width=80)
        self.var_dst = ctk.StringVar(value="B")
        self._campo_txt("Coluna Destino", 5, self.var_dst, width=80)
        ctk.CTkButton(self, text="Processar", command=self._run).grid(row=10, column=0, pady=10, padx=4, sticky="ew")
        ctk.CTkButton(self, text="Sair", fg_color="gray", command=self.destroy).grid(row=10, column=1, pady=10, padx=4, sticky="ew")
        self.log = ctk.CTkTextbox(self, width=600, height=180, state="disabled")
        self.log.grid(row=11, column=0, columnspan=3, pady=(5,10), padx=4)

    def _campo_arquivo(self, rotulo, linha, var):
        ctk.CTkLabel(self, text=rotulo).grid(row=linha, column=0, sticky="w", padx=4, pady=(5,0))
        ctk.CTkEntry(self, textvariable=var, width=420).grid(row=linha, column=1, sticky="ew", padx=4, pady=(5,0))
        ctk.CTkButton(self, text="…", width=30, command=lambda v=var: v.set(filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm")]))).grid(row=linha, column=2, padx=(2,4), pady=(5,0))

    def _campo_txt(self, rotulo, linha, var, width=420):
        ctk.CTkLabel(self, text=rotulo).grid(row=linha, column=0, sticky="w", padx=4, pady=(5,0))
        ctk.CTkEntry(self, textvariable=var, width=width).grid(row=linha, column=1, sticky="w", padx=4, pady=(5,0))

    def _log(self, msg: str):
        self.log.configure(state="normal")
        self.log.insert("end", f"{msg}\n")
        self.log.configure(state="disabled")
        self.log.see("end")
        self.update_idletasks()

    def _exibir_relatorio(self, relatorio: Dict):
        self._log("\n--- RELATÓRIO DE ALTERAÇÕES ---")
        if not any(relatorio.values()): self._log("Nenhuma alteração numérica foi processada."); return
        if alterados := relatorio.get("alterados"):
            self._log(f"\n[VERDE] {len(alterados)} valores foram ALTERADOS:")
            for item in alterados[:10]: self._log(f"  Linha {item['linha']}: {item['ant']} -> {item['atual']}")
            if len(alterados) > 10: self._log("  ...")
        if novos := relatorio.get("novos"):
            self._log(f"\n[AMARELO] {len(novos)} valores NOVOS (eram zero):")
            for item in novos[:10]: self._log(f"  Linha {item['linha']}: {item['atual']}")
            if len(novos) > 10: self._log("  ...")
        if ignorados := relatorio.get("ignorados"):
            self._log(f"\n[NÃO PINTADO] {len(ignorados)} valores IGNORADOS (não encontrados na origem):")
            vals = sorted(list(set(item['valor'] for item in ignorados)))
            self._log(f"  Valores: {vals[:15]}")
            if len(vals) > 15: self._log("  ...")
        self._log("---------------------------------")

    def _run(self):
        self.log.configure(state="normal"); self.log.delete("1.0", "end"); self.log.configure(state="disabled")
        try:
            ori, spr = Path(self.var_ori.get()), Path(self.var_spr.get())
            if not (ori.exists() and spr.exists()): self._log("Por favor, selecione arquivos válidos."); return
            self._log("Iniciando processamento..."); self.update_idletasks()
            relatorio = processar(
                ori=ori, spr=spr, tipo=self.var_tipo.get(), periodo=self.var_per.get(),
                src_txt=self.var_src.get(), dst_txt=self.var_dst.get(),
                start_row=27, dre_start=150, out_dir=None, log=self._log)
            self._log("✔️ Processo finalizado.")
            self._exibir_relatorio(relatorio)
        except Exception as exc:
            logging.exception("Ocorreu um erro no processamento")
            self._log(f"ERRO: {exc}")

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, filename='app_spread.log', filemode='w', format='%(name)s - %(levelname)s - %(message)s')
    app = App()
    app.mainloop()