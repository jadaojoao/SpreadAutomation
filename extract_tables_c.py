"""
app_spread.py
Interface web (Streamlit) para atualizar um arquivo Spread (.xlsx)
a partir de um arquivo de demonstrações financeiras (individual ou
consolidado).  Aceita períodos no formato ano ("2024") ou trimestre
("1T25").  O resultado é disponibilizado para download.

Execute no terminal:
    pip install streamlit pandas openpyxl
    streamlit run extract_tables_c.py
"""

from __future__ import annotations

# ────────────────────────────────────────────────────────────────────────
# Importações padrão + dependências
# ────────────────────────────────────────────────────────────────────────
import io                       # para manter arquivo em memória p/ download
import re                       # expressões regulares – validar período
import tempfile                 # arquivos temporários para os uploads
from pathlib import Path        # manipulação de caminhos de forma segura
from typing import Dict, List, Tuple

import pandas as pd             # tratamento de dados/tabulações
import streamlit as st          # framework web muito simples
from openpyxl import load_workbook  # leitura/escrita de planilhas Excel


# ────────────────────────────────────────────────────────────────────────
# Funções utilitárias
# ────────────────────────────────────────────────────────────────────────
def normaliza_num(valor: object) -> int | None:
    """Remove pontos/virgulas, converte em int. Retorna None se não numérico."""
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None
    if isinstance(valor, (int, float)):
        return int(valor)
    if isinstance(valor, str):
        # Remove separadores de milhar . e , – deixa só dígitos
        limpo: str = valor.strip().replace(".", "").replace(",", "")
        return int(limpo) if limpo.isdigit() else None
    return None


def extrai_formula(expr: str) -> List[int] | None:
    """
    Se a célula for fórmula simples '=1000+2500', devolve [1000, 2500].
    Caso contrário, retorna None.
    """
    if not expr.lstrip().startswith("="):
        return None
    partes: List[int | None] = [normaliza_num(p) for p in expr[1:].split("+")]
    return [p for p in partes if p is not None] if all(partes) else None


def periodos(per_str: str) -> Tuple[str, str, str]:
    """
    Converte '2024'  -> ('2024', '2023', '2022')
            '1T25'   -> ('1T25', '1T24', '1T23')
    """
    per_str = per_str.upper().strip()

    # Caso seja ano cheio
    if re.fullmatch(r"\d{4}", per_str):
        ano = int(per_str)
        return str(ano), str(ano - 1), str(ano - 2)

    # Caso seja trimestre
    m = re.fullmatch(r"([1-4])T(\d{2})", per_str)
    if not m:
        raise ValueError("Período deve ser 'AAAA' ou 'nTAA' (ex.: 2024 ou 1T25).")

    tri, aa = int(m.group(1)), int(m.group(2))

    def fmt(y: int) -> str:
        """Formata trimestre+ano de 2 dígitos."""
        return f"{tri}T{y:02d}"

    return fmt(aa), fmt(aa - 1), fmt(aa - 2)


# ────────────────────────────────────────────────────────────────────────
# Etapa 1 – Filtra, renomeia abas e colunas do arquivo de origem
# ────────────────────────────────────────────────────────────────────────
def prepara_origem(path: Path, tipo: str,
                   atual: str, ant: str, ant2: str) -> Path:
    """
    Cria um novo .xlsx com apenas as abas relevantes e colunas renomeadas.
    Retorna o caminho do novo arquivo.
    """

    # Abas de interesse para cada tipo
    mapa: Dict[str, Dict[str, str]] = {
        "consolidado": {
            "DF Cons Ativo": "cons ativos",
            "DF Cons Passivo": "cons passivos",
            "DF Cons Resultado Periodo": "cons DRE",
            "DF Cons Fluxo de Caixa": "cons DFC",
        },
        "individual": {
            "DF Ind Ativo": "ind ativos",
            "DF Ind Passivo": "ind passivos",
            "DF Ind Resultado Periodo": "ind DRE",
            "DF Ind Fluxo de Caixa": "ind DFC",
        },
    }
    if tipo not in mapa:
        raise ValueError("tipo deve ser 'consolidado' ou 'individual'")

    # Colunas a renomear para cada período
    rename_cols = {
        "Valor Ultimo Exercicio": atual,
        "Valor Penultimo Exercicio": ant,
        "Valor Antepenultimo Exercicio": ant2,
    }

    xls = pd.ExcelFile(path, engine="openpyxl")

    # Caminho do arquivo filtrado generando na mesma pasta
    out_path: Path = path.with_name(f"{path.stem}_{tipo}_{atual}.xlsx")

    # Grava somente as abas selecionadas, já renomeadas
    with pd.ExcelWriter(out_path, engine="openpyxl") as wr:
        for aba_orig, aba_new in mapa[tipo].items():
            if aba_orig in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=aba_orig, engine="openpyxl")
                df = df.rename(columns=lambda c: rename_cols.get(c, c))
                df.to_excel(wr, sheet_name=aba_new, index=False)

    return out_path


# ────────────────────────────────────────────────────────────────────────
# Etapa 2 – Busca valor correspondente linha-a-linha
# ────────────────────────────────────────────────────────────────────────
def valor_corresp(abas: Dict[str, pd.DataFrame],
                  num_ref: int, col_prev: str, col_curr: str) -> int | None:
    """
    Procura num_ref na coluna col_prev; se achar, devolve valor na mesma
    linha em col_curr.  Retorna None se não encontrar.
    """
    for df in abas.values():
        # Pula aba se não tem as colunas necessárias
        if col_prev not in df.columns or col_curr not in df.columns:
            continue
        # Normaliza cada valor da coluna anterior e compara
        linhas = df[df[col_prev].apply(normaliza_num) == num_ref]
        if not linhas.empty:
            return normaliza_num(linhas[col_curr].iloc[0])
    return None


# ────────────────────────────────────────────────────────────────────────
# Etapa 3 – Atualiza o Spread e devolve novo arquivo
# ────────────────────────────────────────────────────────────────────────
def atualizar_spread(spread: Path, abas_path: Path,
                     col_idx: int, atual: str, ant: str) -> Path:
    """
    Percorre o Spread (coluna index col_idx) e preenche nova coluna com
    valores do período 'atual'.  Salva como '<Spread> atual.xlsx'.
    """
    # Lê todas as abas já filtradas
    abas = pd.read_excel(abas_path, sheet_name=None, engine="openpyxl")

    # Abre Spread via openpyxl p/ manter fórmulas
    wb = load_workbook(spread)
    ws = wb.active
    c_src, c_dst = col_idx + 1, col_idx + 2      # Excel é 1-based

    # Linhas do Excel (começando em 1)
    for row in range(1, ws.max_row + 1):
        valor_celula = ws.cell(row=row, column=c_src).value
        if valor_celula is None:
            continue

        # Caso seja fórmula "=num1+num2+..."
        partes = extrai_formula(str(valor_celula))
        if partes:
            novos = [valor_corresp(abas, p, ant, atual) for p in partes]
            # Só grava se encontrou todos os termos
            if all(novos):
                ws.cell(row=row, column=c_dst).value = \
                    "=" + "+".join(str(v) for v in novos)
            continue

        # Caso seja valor isolado
        num = normaliza_num(valor_celula)
        if num is None:
            continue
        novo = valor_corresp(abas, num, ant, atual)
        if novo is not None:
            ws.cell(row=row, column=c_dst).value = novo

    # Cabeçalho da nova coluna = período atual
    ws.cell(row=1, column=c_dst).value = atual

    # Salva novo arquivo com sufixo período
    out_path: Path = spread.with_name(f"{spread.stem} {atual}{spread.suffix}")
    wb.save(out_path)
    return out_path


# ────────────────────────────────────────────────────────────────────────
# Streamlit – Interface web propriamente dita
# ────────────────────────────────────────────────────────────────────────
st.title("Atualizador de Spread (Excel) - https://www.rad.cvm.gov.br/ENET/frmConsultaExternaCVM.aspx?tipoconsulta=CVM&codigoCVM=21393")

# ▸ Upload arquivos
st.header("1. Envie os arquivos")
up_origem = st.file_uploader("Arquivo de origem (.xlsx)", type="xlsx")
up_spread = st.file_uploader("Arquivo Spread (.xlsx)", type="xlsx")

# ▸ Parâmetros
st.header("2. Parâmetros de processamento")
tipo = st.selectbox("Tipo do demonstrativo", ["consolidado", "individual"])
periodo_in = st.text_input("Período (ex.: 2024 ou 1T25)")
col_idx = st.number_input("Coluna dos valores (A=0, B=1…)", min_value=0, value=0)

# ▸ Botão
if st.button("Executar"):
    try:
        # Verificações básicas
        if not up_origem or not up_spread:
            st.error("Envie os dois arquivos .xlsx antes de executar.")
            st.stop()

        # Converte período para atual / anterior / ante-anterior
        atual, ant, ant2 = periodos(periodo_in)

        # Salva uploads em arquivos temporários
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f_ori, \
             tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f_spr:
            f_ori.write(up_origem.getbuffer())
            f_spr.write(up_spread.getbuffer())
            path_ori, path_spr = Path(f_ori.name), Path(f_spr.name)

        st.info("Processando… aguarde alguns segundos.")

        # 1) filtra/renomeia abas
        origem_filtrada = prepara_origem(path_ori, tipo, atual, ant, ant2)
        # 2) atualiza Spread
        novo_spread = atualizar_spread(path_spr, origem_filtrada,
                                       int(col_idx), atual, ant)

        # Exibe botão para baixar o resultado
        with open(novo_spread, "rb") as f_out:
            st.success("✅ Arquivo gerado com sucesso!")
            st.download_button(
                label="📥 Baixar Spread atualizado",
                data=f_out.read(),
                file_name=novo_spread.name,
                mime="application/vnd.openxmlformats-officedocument."
                     "spreadsheetml.sheet",
            )

    except Exception as exc:
        # Captura erros e exibe ao usuário
        st.error(f"Erro: {exc}")
