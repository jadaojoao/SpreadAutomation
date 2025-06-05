import re
from typing import Dict, Union
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def get_table_type_info() -> Dict[str, str]:
    print("Selecione o tipo de tabela:")
    print("1. Consolidado e Controladora")
    print("2. Ativos e passivos: mesma página ao lado")
    print("3. Outro")

    choice = input("Digite o número correspondente ao tipo de tabela: ").strip()
    result = {}

    if choice == "1":
        result["table_type"] = "Consolidado e Controladora"
        valid_options = {"consolidado", "controladora"}

        print("Informe se as duas primeiras colunas são 'Consolidado' ou 'Controladora'.")
        first_col = input("Primeira coluna: ").strip().lower()
        second_col = input("Segunda coluna: ").strip().lower()

        if first_col not in valid_options or second_col not in valid_options:
            raise ValueError("Colunas inválidas. Use apenas 'Consolidado' ou 'Controladora'.")

        result["first_column"] = first_col.capitalize()
        result["second_column"] = second_col.capitalize()

    elif choice == "2":
        result["table_type"] = "Ativos e passivos: mesma página ao lado"
        result["first_column"] = ""
        result["second_column"] = ""

    else:
        result["table_type"] = "Outro"
        result["first_column"] = ""
        result["second_column"] = ""

    return result


def is_date_string(value: Union[str, int, float]) -> bool:
    if not isinstance(value, str):
        value = str(value)
    return bool(
        re.fullmatch(r"\d{4}", value)
        or re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", value)
        or re.fullmatch(r"\d{2}/\d{2}/\d{4}", value)
    )


def remove_empty_rows_and_columns(sheet: Worksheet) -> None:
    for row in reversed(range(1, sheet.max_row + 1)):
        if all(sheet.cell(row=row, column=col).value in (None, "")
               for col in range(1, sheet.max_column + 1)):
            sheet.delete_rows(row)

    for col in reversed(range(1, sheet.max_column + 1)):
        if all(sheet.cell(row=row, column=col).value in (None, "")
               for row in range(1, sheet.max_row + 1)):
            sheet.delete_cols(col)


    max_row = sheet.max_row
    col_idx = 1

    while col_idx <= sheet.max_column:
        split_rows = {}
        for row in range(1, max_row + 1):
            cell = sheet.cell(row=row, column=col_idx)

        if split_rows:
            for row in range(1, max_row + 1):
                for c in range(sheet.max_column, col_idx, -1):
                    sheet.cell(row=row, column=c + 1).value = sheet.cell(row=row, column=c).value

            for row, (word1, word2) in split_rows.items():
                sheet.cell(row=row, column=col_idx).value = word1
                sheet.cell(row=row, column=col_idx + 1).value = word2

            col_idx += 2
        else:
            col_idx += 1


def move_columns_after_break(sheet: Worksheet) -> None:
    second_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    total_cols = len(second_row)

    date_sequence_started = False
    breaking_point = None

    for col_idx, val in enumerate(second_row):
        if is_date_string(val):
            date_sequence_started = True
        elif date_sequence_started and val:
            breaking_point = col_idx
            break

    if breaking_point is None:
        return

    cols_to_move = list(range(breaking_point, total_cols))
    num_rows = sheet.max_row

    for row_idx in range(1, num_rows + 1):
        for i, col_idx in enumerate(cols_to_move):
            value = sheet.cell(row=row_idx, column=col_idx + 1).value
            sheet.cell(row=num_rows + row_idx, column=i + 1).value = value
            sheet.cell(row=row_idx, column=col_idx + 1).value = None


def label_top_of_date_columns_all_sheets(
    file_path: str,
    output_path: str,
    table_type: str,
    first_col_label: str = "",
    second_col_label: str = ""
) -> None:
    wb = load_workbook(file_path)

    for sheet in wb.worksheets:

        if table_type == "Ativos e passivos: mesma página ao lado":
            move_columns_after_break(sheet)

        elif table_type == "Consolidado e Controladora":
            second_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
            date_columns = [
                col_idx for col_idx, cell in enumerate(second_row)
                if cell and is_date_string(cell)
            ]
            for i, col_idx in enumerate(date_columns[:4]):
                label = first_col_label if i < 2 else second_col_label
                col_letter = get_column_letter(col_idx + 1)
                sheet[f"{col_letter}1"] = label

        remove_empty_rows_and_columns(sheet)

    wb.save(output_path)


if __name__ == "__main__":
    try:
        table_info = get_table_type_info()

        label_top_of_date_columns_all_sheets(
            file_path="vibraITR_pages_3_to_4.xlsx",
            output_path="vibraITR_pages_3_to_4_modified.xlsx",
            table_type=table_info["table_type"],
            first_col_label=table_info["first_column"],
            second_col_label=table_info["second_column"]
        )

        print("Arquivo processado com sucesso!")

    except ValueError as e:
        print(f"Erro: {e}")
