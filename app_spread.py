# app_spread.py · v9 (auto-ajuste de colunas Origem/Destino via Período)
# --------------------------------------------------------------------
# • Mapeamento correto das abas (DF Cons/Ind + DMPL, Atual/Último conforme trimestre/ano)
# • Normalização de cabeçalhos distintos para ano × trimestre (Ativo/Passivo, DRE, DFC, DMPL)
# • Varredura do Spread limitada apenas a BP e DRE; DFC e DMPL são tratados em funções específicas
# • Pulagem automática das linhas 199, 209, 210 e 214 na varredura padrão
# • DRE trimestral mapeado manualmente via DRE_MAP
# • Depreciação/Amortização negativa na DFC (agrega múltiplas linhas)
# • Captura de Dividendos e JCP na DMPL:
#     – soma NEGATIVA em linha 210
#     – soma POSITIVA em linha 209
# • Captura de Aumentos de Capital na DMPL:
#     – soma (positiva) em linha 214
# • Ajuste automático de colunas Origem e Destino pelo período:
#     – Ano completo: Origem = H, Destino = J
#     – Trimestre:    Origem = J, Destino = L
# • Atualização ao vivo com xlwings; fallback em openpyxl
# • Destaque em verde (used_vals) para células usadas na atualização
# • Destaque em amarelo para todos os valores que apareceram no Spread
# • Destaque em azul-claro para novidades (prev=0 → atual≠0)
# --------------------------------------------------------------------

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Bootstrap de dependências: instala só se não estiver presente

from __future__ import annotations

import importlib, subprocess, sys

def ensure(pkg: str, module: str | None = None) -> None:
    """
    Garante que o módulo está importável; se não, instala via pip.
    - pkg: nome do pacote no PyPI
    - module: nome do módulo para importar (se diferente de pkg)
    """
    module = module or pkg
    try:
        importlib.import_module(module)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])

# Instalações condicionais
ensure("pandas", "pandas")
ensure("openpyxl", "openpyxl")
ensure("xlwings", "xlwings")
ensure("customtkinter", "customtkinter")

import logging
import re
from pathlib import Path
from typing import Callable, Dict, List, Set, Tuple

import customtkinter as ctk
import pandas as pd
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import (
    column_index_from_string as col2idx,
    get_column_letter as idx2col,
)
try:
    import xlwings as xw
    XLWINGS = True
except ImportError:
    XLWINGS = False


def normaliza_num(v) -> int | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (int, float)):
        return int(v)
    if isinstance(v, str):
        s = v.strip().replace(".", "").replace(",", "")
        if re.fullmatch(r"-?\d+", s):
            return int(s)
    return None


def periodos(p: str) -> Tuple[str, str, str, bool]:
    p = p.upper().strip()
    if re.fullmatch(r"\d{4}", p):
        a = int(p)
        return str(a), str(a-1), str(a-2), False
    m = re.fullmatch(r"([1-4])T(\d{2})", p)
    if not m:
        raise ValueError("Período deve ser AAAA ou nTAA (ex.: 2024 ou 1T25).")
    tri, aa = int(m.group(1)), int(m.group(2))
    f = lambda y: f"{tri}T{y:02d}"
    return f(aa), f(aa-1), f(aa-2), True


def col_txt_to_idx(txt: str) -> int:
    t = txt.strip().upper()
    if t.isdigit():
        return int(t)
    return col2idx(t) - 1


def prepara_origem(
    path: Path,
    tipo: str,
    atual: str,
    ant: str,
    ant2: str,
    is_trim: bool,
    out_dir: Path | None,
) -> Path:
    dst_dir = out_dir or path.parent
    dst_dir.mkdir(parents=True, exist_ok=True)
    dst_path = dst_dir / f"{path.stem}_tratado{path.suffix}"

    chapa = "Cons" if tipo=="consolidado" else "Ind"
    aba_dm = f"DF {chapa} DMPL {'Atual' if is_trim else 'Ultimo'}"
    sheet_map = {
        "consolidado": {
            "DF Cons Ativo": "cons ativos",
            "DF Cons Passivo": "cons passivos",
            "DF Cons Resultado Periodo": "cons DRE",
            "DF Cons Fluxo de Caixa": "cons DFC",
            aba_dm: "cons DMPL",
        },
        "individual": {
            "DF Ind Ativo": "ind ativos",
            "DF Ind Passivo": "ind passivos",
            "DF Ind Resultado Periodo": "ind DRE",
            "DF Ind Fluxo de Caixa": "ind DFC",
            aba_dm: "ind DMPL",
        },
    }[tipo]

    H_ANO = (
        "valor ultimo exercicio",
        "valor penultimo exercicio",
        "valor antepenultimo exercicio",
    )
    H_TRI_AP = ("valor trimestre atual", "valor exercicio anterior")
    H_TRI_RES = (
        "valor acumulado atual exercicio",
        "valor acumulado exercicio anterior",
    )

    def ren_factory(sheet_orig: str) -> Callable[[str], str]:
        low = sheet_orig.lower()
        is_ap = any(k in low for k in ("ativo", "passivo"))
        is_res = "resultado" in low or (is_trim and "fluxo" in low)
        def ren(col: str) -> str:
            c = col.lower().strip()
            if is_trim and is_ap:
                if c.startswith(H_TRI_AP[0]): return atual
                if c.startswith(H_TRI_AP[1]): return ant
            if is_trim and is_res:
                if c.startswith(H_TRI_RES[0]): return atual
                if c.startswith(H_TRI_RES[1]): return ant
            if c.startswith(H_ANO[0]): return atual
            if c.startswith(H_ANO[1]): return ant
            if c.startswith(H_ANO[2]): return ant2
            return col
        return ren

    engine = "openpyxl" if path.suffix.lower() in (".xlsx",".xlsm") else None
    xls = pd.ExcelFile(path, engine=engine)
    with pd.ExcelWriter(dst_path, engine="openpyxl") as wr:
        for orig, novo in sheet_map.items():
            if orig not in xls.sheet_names:
                continue
            df = pd.read_excel(xls, sheet_name=orig, engine=engine)
            # renomeia colunas
            df = df.rename(columns=ren_factory(orig))

            # remove linhas onde todos os períodos são zero
            # coluna de períodos: [atual, ant] em trimestre, ou [atual, ant, ant2] em ano
            period_cols = [atual, ant] if is_trim else [atual, ant, ant2]
            # filtra apenas colunas existentes
            period_cols = [c for c in period_cols if c in df.columns]
            if period_cols:
                # marca linhas onde todos os valores normalizados == 0
                mask_all_zero = df[period_cols] \
                    .applymap(lambda v: normaliza_num(v) == 0) \
                    .all(axis=1)
                df = df.loc[~mask_all_zero]

            df.to_excel(wr, sheet_name=novo, index=False)

    # garante que a primeira aba fique visível
    wb = load_workbook(dst_path)
    wb[wb.sheetnames[0]].sheet_state = "visible"
    wb.save(dst_path)
    return dst_path


def shift_formula(f: str, delta: int) -> str:
    pat = re.compile(
        r"(?<![A-Za-z0-9_])(?:'[^']+'|[A-Za-z0-9_]+)?!"
        r"|(?<![A-Za-z0-9_])(\$?)([A-Za-z]{1,3})(?=\$?\d|:)",
        flags=re.I,
    )
    def repl(m):
        if m.group(1) is None:
            return m.group(0)
        abs_, col = m.group(1), m.group(2)
        try:
            return f"{abs_}{idx2col(col2idx(col)+delta)}"
        except:
            return col
    return pat.sub(repl, f)


def adjust_complex_formula(
    formula: str, delta: int,
    map_number: Callable[[int], int|None],
    used_vals: Set[int]|None=None
) -> str:
    num_pat = re.compile(r"(?<![A-Za-z0-9_])[-+]?\d[\d\.,]*(?![A-Za-z0-9_])")
    f2 = shift_formula(formula, delta)

    def repl(m: re.Match) -> str:
        n = normaliza_num(m.group(0))
        novo = map_number(n)
        if novo is not None and used_vals is not None:
            used_vals.add(novo)
            return str(novo)
        return m.group(0)
    return num_pat.sub(repl, f2)


def valor_corresp(
    abas: Dict[str,pd.DataFrame],
    n: int, prev: str, curr: str
) -> int|None:
    for df in abas.values():
        if prev not in df.columns or curr not in df.columns:
            continue
        hit = df[df[prev].apply(normaliza_num)==n]
        if not hit.empty:
            return normaliza_num(hit[curr].iloc[0])
    return None


def atualizar_ws(
    ws,
    get_val: Callable[[int, int], object],
    set_val: Callable[[int, int, object], None],
    abas: Dict[str, pd.DataFrame],
    src_idx: int,
    dst_idx: int,
    atual: str,
    ant: str,
    start_row: int,
) -> tuple[List[int], Set[int], Set[int]]:
    """
    Copia/ajusta valores/fórmulas da coluna origem→destino, EXCETO
    nas linhas 199, 209, 210 e 214 (tratadas por DFC/DMPL),
    e interrompe em 252.

    - Linhas onde a origem normaliza para 0 são puladas (como vazias).
    - Retira literais '+0' / '-0' de dentro de fórmulas e apaga '=0'.
    """
    c_src, c_dst = src_idx + 1, dst_idx + 1
    delta = c_dst - c_src

    SKIP = {199, 209, 210, 214}
    END_ROW = 252

    skipped_rows: List[int] = []
    skipped_vals: Set[int] = set()
    used_vals: Set[int] = set()

    num_pat = re.compile(r"[-+]?\d[\d\.,]*")
    try:
        max_row = ws.max_row
    except AttributeError:
        max_row = ws.cells.last_cell.row
    last_row = min(max_row, END_ROW)

    empty_streak = 0
    r = start_row
    while empty_streak < 30 and r <= last_row:
        # pula linhas DFC/DMPL
        if r in SKIP:
            r += 1
            continue

        v = get_val(r, c_src)
        # se for zero puro, trata como vazio
        n0 = normaliza_num(v)
        if n0 == 0:
            empty_streak += 1
            r += 1
            continue

        if v in (None, ""):
            empty_streak += 1
            r += 1
            continue

        # achou algo não-zero
        empty_streak = 0
        destino = v
        wrote = False

        # --- lógica original para fórmulas e números não-zero ---
        if isinstance(v, str) and v.startswith("="):
            if not re.search(r"[A-Za-z]", v[1:]):
                # literais de soma/subtração
                def lit_repl(m: re.Match) -> str:
                    tok = m.group(0)
                    n0 = normaliza_num(tok.lstrip("+-"))
                    n1 = valor_corresp(abas, n0, ant, atual)
                    if n1 is not None:
                        used_vals.add(n1)
                        sign = tok[0] if tok[0] in "+-" else ""
                        return f"{sign}{abs(n1)}"
                    return tok

                destino = "=" + num_pat.sub(lit_repl, v[1:])
                wrote = destino != v
            else:
                # fórmula complexa
                mp = lambda n: valor_corresp(abas, n, ant, atual)
                destino = adjust_complex_formula(v, delta, mp, used_vals)
                wrote = destino != v

        elif (n := normaliza_num(v)) is not None:
            # só mapeia números não-zero
            novo = valor_corresp(abas, n, ant, atual)
            if novo is not None:
                destino, wrote = novo, True
                used_vals.add(novo)
        else:
            wrote = True

        # --- limpeza de zeros remanescentes em fórmulas ---
        if isinstance(destino, str) and destino.startswith("="):
            # tira '+0' / '-0' não ligados a referências
            destino = re.sub(r'(?<=\D)[+\-]0(?!\d)', "", destino)
            # se sobrar '=0' ou '=-0', apaga tudo
            if re.fullmatch(r"=[+\-]?0+", destino):
                destino = ""
                wrote = True

        # grava no Spread
        try:
            set_val(r, c_dst, destino)
        except Exception:
            wrote = False

        if not wrote and normaliza_num(v) not in (None, 0):
            skipped_rows.append(r)
            skipped_vals.add(normaliza_num(v) or 0)

        r += 1

    return skipped_rows, skipped_vals, used_vals




DRE_MAP = {
    0:  "Receita de Venda de Bens e/ou Serviços",
    13: "Custo dos Bens e/ou Serviços Vendidos",
    23: "Despesas Gerais e Administrativas",
    25: "Despesas com Vendas",
    26: "Outras Receitas Operacionais",
    27: "Outras Despesas Operacionais",
    29: "Despesas Financeiras",
    30: "Receitas Financeiras",
    41: "Resultado de Equivalência Patrimonial",
    43: "Imposto de Renda e Contribuição Social sobre o Lucro",
}


def aplicar_dre_manual(
    df_dre: pd.DataFrame, sheet,
    col_dst_1based: int, dre_start: int,
    col_valor: str, is_xlwings: bool
) -> None:
    for offset, desc in DRE_MAP.items():
        linha = dre_start + offset
        try:
            raw = df_dre.loc[
                df_dre["Descricao Conta"].str.strip()==desc,
                col_valor
            ].iloc[0]
        except:
            continue
        val = normaliza_num(raw) if normaliza_num(raw) is not None else raw
        if is_xlwings:
            sheet.cells(linha, col_dst_1based).value = val
        else:
            sheet.cell(linha, col_dst_1based, value=val)


def inserir_depreciacao_dfc(
    df_dfc: pd.DataFrame, sheet,
    col_dst_1based: int, linha: int,
    col_valor: str, is_xlwings: bool
) -> int|None:
    if df_dfc is None or col_valor not in df_dfc.columns:
        return None
    desc = df_dfc["Descricao Conta"].astype(str)
    mask = desc.str.contains("deprecia|amortiza", case=False, na=False)
    nums = [normaliza_num(v) for v in df_dfc.loc[mask, col_valor]]
    nums = [n for n in nums if n is not None]
    if not nums:
        return None
    if len(nums)==1:
        total = -abs(nums[0]); val = total
    else:
        terms = "".join(f"-{abs(n)}" for n in nums)
        total = -sum(abs(n) for n in nums)
        val = f"={terms.lstrip('+')}"
    if is_xlwings:
        sheet.cells(linha, col_dst_1based).value = val
    else:
        sheet.cell(linha, col_dst_1based, value=val)
    return total


def inserir_dividendos_dm(
    df_dm: pd.DataFrame,
    sheet,
    col_dst_1based: int,
    linha_neg: int,
    linha_pos: int,
    is_xlwings: bool
) -> tuple[int | None, int | None]:
    """
    Insere na planilha:
      • em `linha_neg` : soma NEGATIVA de todas as contas que contenham
                         'dividendo' ou 'juros sobre capital próprio'
      • em `linha_pos` : soma POSITIVA dessas mesmas contas (se existirem)
    Retorna (total_negativo, total_positivo).
    """
    if df_dm is None:
        return None, None

    # 1) coluna de descrição
    desc_cols = [c for c in df_dm.columns if re.search(r'desc.*conta', c, flags=re.I)]
    if not desc_cols:
        return None, None
    col_desc = desc_cols[0]

    # 2) coluna de valor de patrimônio
    val_cols = [
        c for c in df_dm.columns
        if re.search(r'patrim[oô]nio.*consolidado', c, flags=re.I)
    ]
    if not val_cols:
        val_cols = [
            c for c in df_dm.columns
            if re.search(r'patrim[oô]nio.*liquido', c, flags=re.I)
        ]
    if not val_cols:
        return None, None
    col_valor = val_cols[0]

    # 3) filtra qualquer linha com 'dividendo' ou 'juros sobre capital próprio'
    pattern = r"dividendo|juros sobre capital próprio"
    mask = df_dm[col_desc].astype(str).str.contains(pattern, case=False, na=False)
    nums = [normaliza_num(v) for v in df_dm.loc[mask, col_valor]]
    nums = [n for n in nums if n is not None]
    if not nums:
        return None, None

    # 4) separa negativos e positivos
    negs = [n for n in nums if n < 0]
    poss = [n for n in nums if n > 0]

    total_neg = None
    total_pos = None

    # 5) monta e insere soma NEGATIVA em `linha_neg`
    if negs:
        if len(negs) == 1:
            total_neg = negs[0]
            val_neg = total_neg
        else:
            terms = "".join(f"-{abs(n)}" for n in negs)
            total_neg = -sum(abs(n) for n in negs)
            val_neg = f"={terms.lstrip('+')}"
        if is_xlwings:
            sheet.cells(linha_neg, col_dst_1based).value = val_neg
        else:
            sheet.cell(row=linha_neg, column=col_dst_1based, value=val_neg)

    # 6) monta e insere soma POSITIVA em `linha_pos`
    if poss:
        total_pos = sum(poss)
        if len(poss) == 1:
            val_pos = total_pos
        else:
            terms = "+".join(str(n) for n in poss)
            val_pos = f"={terms}"
        if is_xlwings:
            sheet.cells(linha_pos, col_dst_1based).value = val_pos
        else:
            sheet.cell(row=linha_pos, column=col_dst_1based, value=val_pos)

    return total_neg, total_pos


def inserir_aumentos_capital_dm(
    df_dm: pd.DataFrame,
    sheet,
    col_dst_1based: int,
    linha: int,
    is_xlwings: bool
) -> int | None:
    """
    Na aba DMPL, filtra todas as linhas que contenham 'Aumento de Capital'
    (ou variações) e soma esses valores (de patrimônio líquido).
    Insere o total na linha `linha` (sempre como positivo ou fórmula =x+y+...).
    Retorna o total (int) para adicionar a used_vals.
    """
    # 1) identifica coluna de descrição
    desc_cols = [c for c in df_dm.columns if re.search(r'desc.*conta', c, flags=re.I)]
    if not desc_cols:
        return None
    col_desc = desc_cols[0]

    # 2) identifica coluna de valor de patrimônio
    val_cols = [c for c in df_dm.columns if re.search(r'patrim[oô]nio.*consolidado', c, flags=re.I)]
    if not val_cols:
        val_cols = [c for c in df_dm.columns if re.search(r'patrim[oô]nio.*liquido', c, flags=re.I)]
    if not val_cols:
        return None
    col_valor = val_cols[0]

    # 3) filtra 'Aumento(s) de Capital'
    mask = df_dm[col_desc].astype(str)\
               .str.contains(r'aumentos?\s+de\s+capital', case=False, na=False)
    series = df_dm.loc[mask, col_valor]
    nums = [normaliza_num(v) for v in series]
    nums = [n for n in nums if n is not None]
    if not nums:
        return None

    # 4) monta soma
    if len(nums) == 1:
        total = nums[0]
        val = total
    else:
        terms = "+".join(str(n) for n in nums)
        total = sum(nums)
        val = f"={terms}"

    # 5) grava no Spread
    if is_xlwings:
        sheet.cells(linha, col_dst_1based).value = val
    else:
        sheet.cell(row=linha, column=col_dst_1based, value=val)

    return total


def destacar_inseridos(
    orig_tratada: Path, used_vals: Set[int],
    atual: str, prefer_xlwings: bool=True
) -> None:
    if not used_vals:
        return
    if prefer_xlwings and XLWINGS:
        try:
            wb = xw.Book(str(orig_tratada))
            for sht in wb.sheets:
                hdrs = sht.range("A1").expand("right").value
                hdrs = hdrs if isinstance(hdrs, list) else [hdrs]
                cols = [i+1 for i,h in enumerate(hdrs) if str(h).strip()==atual]
                last = sht.cells.last_cell.row
                for c in cols:
                    vals = sht.range((2,c),(last,c)).value
                    vals = vals if isinstance(vals,list) else [vals]
                    for i,v in enumerate(vals, start=2):
                        if normaliza_num(v) in used_vals:
                            cell = sht.cells(i,c)
                            cell.color = (204,255,204)
                            cell.api.Font.Bold = True
            wb.save()
            return
        except:
            pass
    wb = load_workbook(orig_tratada, keep_vba=orig_tratada.suffix.lower()==".xlsm")
    fill,bold = PatternFill("solid", fgColor="CCFFCC"), Font(bold=True)
    for ws in wb.worksheets:
        cols = [c.column for c in ws[1] if str(c.value).strip()==atual]
        for row in ws.iter_rows(min_row=2):
            for c in cols:
                cell = row[c-1]
                if normaliza_num(cell.value) in used_vals:
                    cell.fill, cell.font = fill,bold
    wb.save(orig_tratada)


def destacar_novos(
    orig_tratada: Path, prev: str, atual: str,
    prefer_xlwings: bool=True
) -> None:
    from openpyxl.styles import PatternFill, Font
    rgb_fill = PatternFill("solid", fgColor="99CCFF")
    bold = Font(bold=True)
    if prefer_xlwings and XLWINGS:
        try:
            for bk in xw.books:
                if Path(bk.fullname).resolve() == orig_tratada.resolve():
                    wb = bk; break
            else:
                wb = xw.Book(str(orig_tratada))
            for sht in wb.sheets:
                hdrs = sht.range("A1").expand("right").value
                hdrs = hdrs if isinstance(hdrs,list) else [hdrs]
                if prev not in hdrs or atual not in hdrs:
                    continue
                c_prev = hdrs.index(prev)+1
                c_atual = hdrs.index(atual)+1
                last = sht.cells.last_cell.row
                vp = sht.range((2,c_prev),(last,c_prev)).value
                va = sht.range((2,c_atual),(last,c_atual)).value
                vp = vp if isinstance(vp,list) else [vp]
                va = va if isinstance(va,list) else [va]
                for i,(pv,av) in enumerate(zip(vp,va), start=2):
                    if normaliza_num(pv)==0 and normaliza_num(av) not in (None,0):
                        cell = sht.cells(i,c_atual)
                        cell.color = (153,204,255)
                        cell.api.Font.Bold = True
            wb.save()
            return
        except:
            pass
    wb = load_workbook(orig_tratada, keep_vba=orig_tratada.suffix.lower()==".xlsm")
    for ws in wb.worksheets:
        headers = {cell.value:cell.column for cell in ws[1]}
        c_prev = headers.get(prev)
        c_atual= headers.get(atual)
        if not c_prev or not c_atual:
            continue
        for row in ws.iter_rows(min_row=2):
            pv = row[c_prev-1].value
            av = row[c_atual-1].value
            if normaliza_num(pv)==0 and normaliza_num(av) not in (None,0):
                row[c_atual-1].fill = rgb_fill
                row[c_atual-1].font = bold
    wb.save(orig_tratada)


def coletar_vals_do_spread(
    spread_path: Path, dst_idx: int, start_row: int
) -> Set[int]:
    wb = load_workbook(spread_path, data_only=True)
    ws = wb.active
    vals, empty = set(), 0
    r = start_row
    while empty<30 and r<=ws.max_row:
        raw = ws.cell(r, dst_idx+1).value
        if raw in (None,""):
            empty+=1
        else:
            empty=0
            n = normaliza_num(raw)
            if n is not None:
                vals.add(n)
        r+=1
    return vals


def processar(
    ori: Path,
    spr: Path,
    tipo: str,
    periodo: str,
    src_txt: str,
    dst_txt: str,
    start_row: int,
    dre_start: int,
    out_dir: Path | None = None,
    log: Callable[[str], None] = print,
) -> Path:
    """Pipeline principal: gera Spread, origem tratada e destaques."""
    src_idx = col_txt_to_idx(src_txt)
    dst_idx = col_txt_to_idx(dst_txt)
    atual, ant, ant2, is_trim = periodos(periodo)

    # prepara e lê as abas tratadas
    orig_path = prepara_origem(ori, tipo, atual, ant, ant2, is_trim, out_dir)
    abas = pd.read_excel(orig_path, sheet_name=None, engine="openpyxl")

    # extraia só as abas de BP e DRE (usando os nomes que você já mapeou no prepara_origem)
    prefix = 'cons' if tipo=='consolidado' else 'ind'
    orig_abas = {
        f"{prefix} ativos": abas[f"{prefix} ativos"],
        f"{prefix} passivos": abas[f"{prefix} passivos"],
        f"{prefix} DRE":     abas[f"{prefix} DRE"],
    }

    df_dre = abas.get(f"{'cons' if tipo=='consolidado' else 'ind'} DRE")
    df_dfc = abas.get(f"{'cons' if tipo=='consolidado' else 'ind'} DFC")
    df_dm  = abas.get(f"{'cons' if tipo=='consolidado' else 'ind'} DMPL")

    used_vals: Set[int] = set()

    # --- tentativa com xlwings ---
    if XLWINGS and spr.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            # conecta ou abre o spread
            for bk in xw.books:
                if Path(bk.fullname).resolve() == spr.resolve():
                    wb = bk; break
            else:
                wb = xw.Book(str(spr))

            # escolhe a aba certa
            nomes = [s.name for s in wb.sheets]
            sht = wb.sheets["Entrada de Dado"] if "Entrada de Dado" in nomes else wb.sheets.active

            # getters/setters
            get_val = lambda r, c: sht.cells(r, c).formula or sht.cells(r, c).value
            def set_val(r, c, v):
                attr = "formula" if isinstance(v, str) and v.startswith("=") else "value"
                setattr(sht.cells(r, c), attr, v)

            # 1) DADOS PRINCIPAIS
            _, _, used = atualizar_ws(
                sht, get_val, set_val, orig_abas,
                src_idx, dst_idx, atual, ant, start_row
            )
            used_vals |= used

            # 2) DRE MANUAL (trimestral)
            if is_trim and df_dre is not None:
                aplicar_dre_manual(df_dre, sht, dst_idx+1, dre_start, atual, True)

            # 3) DEPRECIAÇÃO/AMORTIZAÇÃO
            if df_dfc is not None:
                if v199 := inserir_depreciacao_dfc(df_dfc, sht, dst_idx+1, 199, atual, True):
                    used_vals.add(v199)

            # 4) DIVIDENDOS DMPL → linhas 210 (negativos) e 209 (positivos)
            if df_dm is not None:
                col_dm = dst_idx + 1
                neg, pos = inserir_dividendos_dm(df_dm, sht, col_dm, 210, 209, True)
                if neg is not None:
                    used_vals.add(neg)
                if pos is not None:
                    used_vals.add(pos)
            
            # --- insere Aumentos de Capital na linha 214 ---
            if df_dm is not None:
                # mesma coluna de destino usada para dividendos
                col_dm = dst_idx + 1
                if v214 := inserir_aumentos_capital_dm(df_dm, sht, col_dm, 214, True):
                    used_vals.add(v214)

            # 5) recalcula e salva
            wb.app.calculate()
            wb.save()

        except Exception as exc:
            log(f"xlwings falhou, usando fallback: {exc}")

    # --- fallback openpyxl ---
    is_xlsm = spr.suffix.lower() == ".xlsm"
    wb2 = load_workbook(spr, keep_vba=is_xlsm)
    ws2 = wb2.active

    _, _, used = atualizar_ws(
        ws2,
        lambda r, c: ws2.cell(r, c).value,
        lambda r, c, v: setattr(ws2.cell(r, c), "value", v),
        orig_abas, src_idx, dst_idx, atual, ant, start_row
    )
    used_vals |= used

    if is_trim and df_dre is not None:
        aplicar_dre_manual(df_dre, ws2, dst_idx+1, dre_start, atual, False)

    if df_dfc is not None:
        if v199 := inserir_depreciacao_dfc(df_dfc, ws2, dst_idx+1, 199, atual, False):
            used_vals.add(v199)

    # DIVIDENDOS DMPL → linhas 210 e 209
    if df_dm is not None:
        col_dm = dst_idx + 1
        neg, pos = inserir_dividendos_dm(df_dm, ws2, col_dm, 210, 209, False)
        if neg is not None:
            used_vals.add(neg)
        if pos is not None:
            used_vals.add(pos)

    # ... depois de inserir_depreciacao_dfc e inserção de dividendos ...
    if df_dm is not None:
        col_dm = dst_idx + 1
        if v214 := inserir_aumentos_capital_dm(df_dm, ws2, col_dm, 214, False):
            used_vals.add(v214)


    out_name = f"{spr.stem} {atual}{'.xlsm' if is_xlsm else '.xlsx'}"
    spr = spr.with_name(out_name)
    wb2.save(spr)

    # --- destaques finais ---
    spread_vals = coletar_vals_do_spread(spr, dst_idx, start_row)
    highlight = used_vals.union(spread_vals)
    destacar_inseridos(orig_path, highlight, atual, prefer_xlwings=XLWINGS)
    destacar_novos(orig_path, ant, atual, prefer_xlwings=XLWINGS)

    log(f"Origem tratada salva em: {orig_path}")
    missing = spread_vals - used_vals
    if missing:
        log(f"⚠️  {len(missing)} valores entraram no Spread mas não estavam em used_vals:")
        log(f"    {sorted(missing)}")
    return spr




class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Atualizador de Spread")
        self.grid_columnconfigure((0,1), weight=1)

        # variáveis
        self.var_ori = ctk.StringVar()
        self.var_spr = ctk.StringVar()
        self.var_tipo = ctk.StringVar(value="consolidado")
        self.var_per = ctk.StringVar()
        self.var_src = ctk.StringVar()
        self.var_dst = ctk.StringVar()

        # campos de arquivo
        self._campo_arquivo("Arquivo Origem", 0, self.var_ori)
        self._campo_arquivo("Arquivo Spread", 1, self.var_spr)

        # tipo consol/indiv
        ctk.CTkLabel(self, text="Tipo").grid(row=2, column=0, sticky="w", padx=4)
        ctk.CTkOptionMenu(self, variable=self.var_tipo,
                          values=["consolidado", "individual"]
                          ).grid(row=2, column=1, sticky="ew", padx=4)

        # período e colunas
        self._campo_txt("Período (Ex: 2024 ou 1T25)", 3, self.var_per)
        self._campo_txt("Coluna Origem",        4, self.var_src, width=80)
        self._campo_txt("Coluna Destino",       5, self.var_dst, width=80)

        # trace para ajustar colunas automaticamente
        self.var_per.trace_add("write", self._on_period_change)

        # botões e log
        ctk.CTkButton(self, text="Processar", command=self._run).grid(row=10, column=0, pady=10, padx=4, sticky="ew")
        ctk.CTkButton(self, text="Sair", fg_color="gray",
                      command=self.destroy
                      ).grid(row=10, column=1, pady=10, padx=4, sticky="ew")
        self.log = ctk.CTkTextbox(self, width=600, height=150, state="disabled")
        self.log.grid(row=11, column=0, columnspan=2, pady=(5,10), padx=4)

        logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

        # inicializa colunas padrão
        self._on_period_change()

    def _on_period_change(self, *args):
        # define origem/destino conforme tipo de período
        p = self.var_per.get().strip()
        try:
            _, _, _, is_trim = periodos(p)
        except Exception:
            return
        if is_trim:
            # trimestre: Origem J, Destino L
            self.var_src.set("J")
            self.var_dst.set("L")
        else:
            # ano completo: Origem H, Destino J
            self.var_src.set("H")
            self.var_dst.set("J")

    def _campo_arquivo(self, rotulo, linha, var):
        ctk.CTkLabel(self,text=rotulo).grid(row=linha,column=0,sticky="w",padx=4)
        ctk.CTkEntry(self,textvariable=var,width=420).grid(row=linha,column=1,sticky="ew",padx=4)
        def escolhe():
            f = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx *.xlsm *.xls")])
            if f: var.set(f)
        ctk.CTkButton(self,text="…",width=30,command=escolhe).grid(row=linha,column=2,padx=2)

    def _campo_txt(self, rotulo, linha, var, width=420):
        ctk.CTkLabel(self,text=rotulo).grid(row=linha,column=0,sticky="w",padx=4)
        ctk.CTkEntry(self,textvariable=var,width=width).grid(row=linha,column=1,sticky="w",padx=4)

    def _log(self,msg):
        self.log.configure(state="normal")
        self.log.insert("end",msg+"\n")
        self.log.configure(state="disabled")
        self.log.see("end")

    def _run(self):
        try:
            ori = Path(self.var_ori.get())
            spr = Path(self.var_spr.get())
            if not (ori.exists() and spr.exists()):
                self._log("Selecione arquivos válidos."); return
            out = processar(
                ori=ori, spr=spr, tipo=self.var_tipo.get(),
                periodo=self.var_per.get(),
                src_txt=self.var_src.get(), dst_txt=self.var_dst.get(),
                start_row=27, dre_start=150, out_dir=None,
                log=self._log
            )
            self._log(f"✔️ Finalizado: {out}")
        except Exception as e:
            logging.exception("Erro no processamento")
            self._log(f"Erro: {e}")


if __name__ == "__main__":
    App().mainloop()
